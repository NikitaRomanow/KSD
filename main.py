
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

path = str(input("Укажите путь к папке в которой находятся файлы с КСД: "))
#path = "C:/Users/nromanov/Desktop/Оренбург_КСД"
k=3
count = 0
new_wb = load_workbook("Анализ_КСД.xlsx")
new_sheet = new_wb.worksheets[0]
new_sheet_2 = new_wb.worksheets[1]

for root,dirs,files in os.walk(path):
    files_count = len(files)-2
    for file in files:
        if file.endswith(".xlsx") and file != "Анализ_КСД.xlsx":
            filename_parse = file.split("_")
            new_parse = filename_parse[0].split(" ")
            if ("Pзаб стат" not in filename_parse) and ("Pзаб статика" not in filename_parse) and ("стат" not in filename_parse)\
                    and ("КВД" not in filename_parse) and ("не проход" not in filename_parse):
                try:
                    ALLOWED_NAMES = ["Рзаб дин","Рзаб динамика","Рзаб"]
                    if (filename_parse[1] in ALLOWED_NAMES) or (new_parse[1] in ALLOWED_NAMES):
                        count += 1
                        print(file)
                        print(f"Выполняется: {count} из {files_count} КСД")

                        wellname = new_parse[0]
                        data = ".".join(filename_parse[-3:])
                        data_split = data.split(".")
                        new_data = ".".join(data_split[:3])

                        #try: #проверка для избежания ошибок
                        wb = load_workbook(file, data_only=True)
                        sheet_shapka = wb.worksheets[0]
                        sheet = wb.worksheets[2]
                        sheet_dannie = wb.worksheets[5]

                        new_sheet.cell(row=k, column=1).value = file  # имя файла
                        new_sheet.cell(row=k, column=2).value = wellname  # номер скважины
                        new_sheet.cell(row=k, column=3).value = new_data  # дата замера
                        new_wb.save("Анализ_КСД.xlsx")

                        #считывание информации из шапки
                        #try:
                        sh_row = 10
                        a=2
                        if sheet_shapka.cell(row=sh_row, column=1).value != None:
                            pass
                        else:
                            sh_row += 1

                        a=7

                        p_skv = str(sheet_shapka.cell(row=sh_row + 6, column=a).value) #Pскважины
                        if sheet_shapka.cell(row=sh_row + 6, column=a).value == None:
                            p_skv = str(sheet_shapka.cell(row=sh_row + 5, column=a + 1).value)
                        new_sheet.cell(row=k, column=6).value = p_skv

                        if (sheet_shapka.cell(row=sh_row, column=a+2).value == "Р на кровлю, атм")\
                                or (sheet_shapka.cell(row=sh_row, column=a+2).value == "               Р на кровлю, атм")\
                                or (sheet_shapka.cell(row=sh_row, column=a+2).value == "Р на кровлю,ат"):
                            a+=2

                        p_trub = sheet_shapka.cell(row=sh_row + 6, column=a+2).value #Pтрубное
                        if sheet_shapka.cell(row=sh_row + 6, column=a+2).value == None:
                            p_trub = sheet_shapka.cell(row=sh_row + 5, column=a + 3).value
                        new_sheet.cell(row=k, column=5).value = p_trub

                        p_zatrub = sheet_shapka.cell(row=sh_row + 6, column=a + 4).value #Pзатрубное
                        if sheet_shapka.cell(row=sh_row + 6, column=a + 4).value == None:
                            p_zatrub = sheet_shapka.cell(row=sh_row + 5, column=a + 5).value
                        new_sheet.cell(row=k, column=8).value = p_zatrub

                        sh_temp = sheet_shapka.cell(row=sh_row + 6, column=a + 9).value  # Температура
                        if sheet_shapka.cell(row=sh_row + 6, column=a + 9).value == None:
                            sh_temp = sheet_shapka.cell(row=sh_row + 2, column=a + 8).value
                        new_sheet.cell(row=k, column=4).value = sh_temp

                        exploration_metod = str(sheet_shapka.cell(row=5, column=5).value)
                        new_sheet.cell(row=k, column=20).value = exploration_metod


                        #except:
                            #pass

                        # обработка папки технолога

                        i = 8
                        while sheet.cell(row=i, column=2).value != None: #пропускаем таблицу со спуском
                            i +=1
                        i +=5

                        # определение начала и конца таблицы
                        start = i

                        while sheet.cell(row=start + 1, column=2).value != None:
                            start += 1
                        end = start
                        start = i

                        start_pres = sheet.cell(row=start, column=2).value
                        end_depth = sheet.cell(row=end, column=1).value
                        end_pres = sheet.cell(row=end, column=2).value
                        end_temp = sheet.cell(row=end, column=3).value
                        end_grav = sheet.cell(row=end, column=4).value

                        j = 0

                        if end_grav>850: #проверка скопления воды
                            new_sheet.cell(row=k, column=14).value = "Есть"
                            new_wb.save("Анализ_КСД.xlsx")
                            grav = float(sheet.cell(row=end, column=4).value)
                            while grav>850 and ((end - j) > start):  # проверяем скопление воды
                                j +=1
                                grav = float(sheet.cell(row=end - j, column=4).value)
                            else:
                                no_water_depth = float(sheet.cell(row=end - j, column=1).value)
                                no_water_pres = sheet.cell(row=end-j, column=2).value
                                no_water_temp = sheet.cell(row=end-j, column=3).value
                                no_water_grav = sheet.cell(row=end-j, column=4).value
                        else:
                            no_water_depth = end_depth
                            no_water_pres = end_pres
                            no_water_grav = end_grav
                            no_water_temp = end_temp

                        if exploration_metod == "Газлифтный" or exploration_metod == "Газлифт"\
                                or exploration_metod == " Газлифтный" or exploration_metod == " всг" or exploration_metod == " Газлифт":  # газлифт

                            l = 2

                            while sheet.cell(row=i+1, column=2).value != None: #проверяем отклонение по давлению

                                pres = float(sheet.cell(row=i, column=2).value)
                                new_pres = float(sheet.cell(row=i + 1, column=2).value)
                                pres_diff = float(new_pres - pres)
                                grav = float(sheet.cell(row=i, column=4).value)
                                grad_temp = float(sheet.cell(row=i, column=6).value)
                                new_grav = float(sheet.cell(row=i + 1, column=4).value)
                                new_grad_temp = float(sheet.cell(row=i + 1, column=6).value)
                                grav_diff = float(new_grav - grav)
                                temp_diff = float(grad_temp - new_grad_temp)
                                depth = float(sheet.cell(row=i, column=1).value)
                                new_depth = float(sheet.cell(row=i+1, column=1).value)
                                temp = float(sheet.cell(row=i, column=3).value)
                                new_temp = float(sheet.cell(row=i+1, column=3).value)

                                if pres_diff>4 and grav<850 and new_grav<850: #фиксирование перепада по давлению
                                    b=i
                                    while b<end and b<(i+4): #проверка перепада давления для следующих 4 строк
                                        b_pres = float(sheet.cell(row=b, column=2).value)
                                        b_new_pres = float(sheet.cell(row=b + 1, column=2).value)
                                        b_pres_diff = float(b_new_pres - b_pres)
                                        if b_pres_diff<1.5:
                                            break
                                        else:
                                            b+=1
                                    else:
                                        #проверяем глубину для клапанов
                                        NO_GAUGE_NEARBY = False

                                        m = 2
                                        while new_sheet_2.cell(row=l,column=1).value != None and not NO_GAUGE_NEARBY :  # проверка номера скважины в листе
                                            if str(new_sheet_2.cell(row=l, column=1).value) == str(wellname):  # если скважина есть
                                                while new_sheet_2.cell(row=l,column=m).value != None:  # проверяем глубины клапанов
                                                    #try:
                                                    if abs(float(new_sheet_2.cell(row=l, column=m).value) - new_depth) < 150:
                                                        new_sheet.cell(row=k, column=16).value = new_sheet_2.cell(row=l,column=m).value
                                                        new_wb.save("Анализ_КСД.xlsx")
                                                        if new_sheet_2.cell(row=l, column=m + 1).value != None:
                                                            new_sheet.cell(row=k, column=16).fill = PatternFill(start_color="ffff00", end_color="ffff00",fill_type="solid")
                                                            new_wb.save("Анализ_КСД.xlsx")
                                                            #если перепад в зоне клапана и этот клапан не ЦКСОК
                                                            try:
                                                                grad_pres_before = (pres - start_pres) / (i - start)  # градиент до
                                                                new_sheet.cell(row=k,column=17).value = grad_pres_before
                                                                new_wb.save("Анализ_КСД.xlsx")
                                                            except ZeroDivisionError:
                                                                new_sheet.cell(row=k,column=17).value = "Перепад не определен"
                                                                new_wb.save("Анализ_КСД.xlsx")
                                                            try:
                                                                grad_pres_after = (end_pres - pres) / (end - i)  # градиент после
                                                                new_sheet.cell(row=k, column=18).value = grad_pres_after
                                                                new_wb.save("Анализ_КСД.xlsx")
                                                            except ZeroDivisionError:
                                                                new_sheet.cell(row=k,column=18).value = "Перепад не определен"
                                                                new_wb.save("Анализ_КСД.xlsx")
                                                            if ((new_grav > (grav * 2)) or (new_grav > (grav + 200))) and temp_diff > 0.03:
                                                                new_sheet.cell(row=k, column=9).value = depth  # глубина
                                                                new_sheet.cell(row=k,column=10).value = pres  # давление
                                                                new_sheet.cell(row=k,column=11).value = grav  # плотность
                                                                new_sheet.cell(row=k,column=12).value = temp  # температура
                                                                new_sheet.cell(row=k,column=13).value = "Пропуск клапана по всем показателям"  # комментарий
                                                                new_wb.save("Анализ_КСД.xlsx")

                                                                break
                                                            elif (new_grav > (grav * 2)) or (new_grav > (grav + 200)):
                                                                new_sheet.cell(row=k, column=9).value = depth  # глубина
                                                                new_sheet.cell(row=k,column=10).value = pres  # давление
                                                                new_sheet.cell(row=k,column=11).value = grav  # плотность
                                                                new_sheet.cell(row=k,column=12).value = temp  # температура
                                                                new_sheet.cell(row=k,column=13).value = "Пропуск клапана по давлению и плотности"  # комментарий
                                                                new_wb.save("Анализ_КСД.xlsx")

                                                                break
                                                            else:
                                                                new_sheet.cell(row=k, column=9).value = depth  # глубина
                                                                new_sheet.cell(row=k,column=10).value = pres  # давление
                                                                new_sheet.cell(row=k,column=11).value = grav  # плотность
                                                                new_sheet.cell(row=k,column=12).value = temp  # температура
                                                                new_sheet.cell(row=k,column=13).value = "Необъяснимый скачок давления в районе клапана"  # комментарий
                                                                new_wb.save("Анализ_КСД.xlsx")

                                                                break
                                                        else: #ЦКСОК
                                                            if end_grav<850:
                                                                new_sheet.cell(row=k, column=9).value = end_depth  # глубина
                                                                new_sheet.cell(row=k,column=10).value = end_pres  # давление
                                                                new_sheet.cell(row=k,column=11).value = end_grav  # плотность
                                                                new_sheet.cell(row=k,column=12).value = end_temp  # температура
                                                                new_sheet.cell(row=k,column=13).value = "Перепад давления в ЦКСОК. Замер взят в нижней точке"
                                                                new_wb.save("Анализ_КСД.xlsx")

                                                                break

                                                    m += 1
                                                    # except:
                                                    #     new_sheet.cell(row=k, column=16).value = "Произошла ошибка"
                                                    #     new_wb.save("Анализ_КСД.xlsx")
                                                    #     break
                                                else:
                                                    NO_GAUGE_NEARBY = True

                                            l += 1

                                elif depth == no_water_depth and new_sheet.cell(row=k, column=9).value == None:

                                    new_sheet.cell(row=k, column=9).value = no_water_depth  # глубина
                                    new_sheet.cell(row=k, column=10).value = no_water_pres  # давление
                                    new_sheet.cell(row=k, column=11).value = no_water_grav  # плотность
                                    new_sheet.cell(row=k, column=12).value = no_water_temp  # температура
                                    new_sheet.cell(row=k, column=13).value = "Скопление воды, замер взят выше точки скопления воды"  # комментарий
                                    new_wb.save("Анализ_КСД.xlsx")
                                    break

                                i +=1

                            #если пропусков не обнаружено
                            if new_sheet.cell(row=k, column=9).value == None:
                                new_sheet.cell(row=k, column=1).value = file  # имя файла
                                new_sheet.cell(row=k, column=2).value = wellname  # номер скважины
                                new_sheet.cell(row=k, column=3).value = new_data  # дата замера
                                new_sheet.cell(row=k, column=9).value = end_depth  # глубина
                                new_sheet.cell(row=k, column=10).value = end_pres  # давление
                                new_sheet.cell(row=k, column=11).value = end_grav  # плотность
                                new_sheet.cell(row=k, column=12).value = end_temp  # температура
                                new_sheet.cell(row=k,column=13).value = "Пропуск не обнаружен. Замер взят в нижней точке"  # комментарий
                                new_wb.save("Анализ_КСД.xlsx")
                        elif exploration_metod == "ЭЦН" or exploration_metod == "Эцн" or exploration_metod == "эцн"\
                                or exploration_metod == " ЭЦН" or exploration_metod == " эцн":

                            new_sheet.cell(row=k, column=9).value = end_depth  # глубина
                            new_sheet.cell(row=k, column=10).value = end_pres  # давление
                            new_sheet.cell(row=k, column=11).value = end_grav  # плотность
                            new_sheet.cell(row=k, column=12).value = end_temp #температура
                            new_sheet.cell(row=k, column=13).value = "Замер взят в нижней точке"  # комментарий
                            new_wb.save("Анализ_КСД.xlsx")

                        elif exploration_metod == "Фонтан" or exploration_metod == "Фонтанный" or exploration_metod == "фонтанный" or\
                            exploration_metod == "фонтан" or exploration_metod == "фонт" or exploration_metod == " фонт" or exploration_metod == "НЕФТЯНАЯ":

                            if no_water_depth == end_depth:
                                new_sheet.cell(row=k, column=9).value = end_depth  # глубина
                                new_sheet.cell(row=k, column=10).value = end_pres  # давление
                                new_sheet.cell(row=k, column=11).value = end_grav  # плотность
                                new_sheet.cell(row=k, column=12).value = end_temp  # температура
                                new_sheet.cell(row=k, column=13).value = "Замер взят в нижней точке"  # комментарий
                                new_wb.save("Анализ_КСД.xlsx")

                            else:
                                new_sheet.cell(row=k, column=9).value = no_water_depth  # глубина
                                new_sheet.cell(row=k, column=10).value = no_water_pres  # давление
                                new_sheet.cell(row=k, column=11).value = no_water_grav  # плотность
                                new_sheet.cell(row=k, column=12).value = no_water_temp  # температура
                                new_sheet.cell(row=k, column=13).value = "Замер взят над скоплением воды"  # комментарий
                                new_wb.save("Анализ_КСД.xlsx")



                        #проверка стабильности замера забойного давления

                        c = 30

                        WRONG_PRESSURE = False

                        while sheet_dannie.cell(row=c, column=4).value != None and not WRONG_PRESSURE:
                            md = sheet_dannie.cell(row=c, column=4).value
                            if md < 1000:
                                c+=1
                            else:
                                if sheet_dannie.cell(row=c+1, column=4).value==md and sheet_dannie.cell(row=c+2, column=4).value==md: #проверка одинкаовой глубины
                                    dannie_pzab = sheet_dannie.cell(row=c, column=2).value
                                    while sheet_dannie.cell(row=c+1, column=4).value ==md:

                                        next_dannie_pzab = sheet_dannie.cell(row=c+1, column=2).value
                                        dannie_pres_diff = abs(float(next_dannie_pzab-dannie_pzab))
                                        WRONG_PRESSURE = dannie_pres_diff>1
                                        if WRONG_PRESSURE:
                                            new_sheet.cell(row=k, column=19).value = "Невалидный замер Pзаб"
                                            new_wb.save("Анализ_КСД.xlsx")
                                            break
                                        else:
                                            c+=1
                                    else:
                                        new_sheet.cell(row=k, column=19).value = "Валидный замер Pзаб"
                                        new_wb.save("Анализ_КСД.xlsx")
                                        break
                                else:
                                    c+=1

                        k += 1
                        #except: #если ошибка случилась

                            # new_sheet.cell(row=k, column=1).value = file  # имя файла
                            # new_sheet.cell(row=k, column=2).value = wellname  # номер скважины
                            # new_sheet.cell(row=k, column=3).value = new_data  # дата замера
                            # new_sheet.cell(row=k, column=13).value = "Произошла ошибка"
                            # k+=1
                            # new_wb.save("Анализ_КСД.xlsx")

                except:
                    new_sheet.cell(row=k, column=1).value = file  # имя файла
                    new_sheet.cell(row=k, column=13).value = "Произошла ошибка"
                    k+=1
                    new_wb.save("Анализ_КСД.xlsx")

            else:
                pass